import io
import json
import os
import uuid
from datetime import datetime, date
from functools import wraps
from urllib.parse import urlparse

import bcrypt
from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, send_file, jsonify, g, abort,
)
from flask_login import (
    LoginManager, UserMixin, login_user, logout_user,
    login_required, current_user,
)
from flask_wtf.csrf import CSRFProtect
from werkzeug.utils import secure_filename

from config import Config
from db import get_db, init_db

app = Flask(__name__)
app.config.from_object(Config)

csrf = CSRFProtect(app)

login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Silakan login terlebih dahulu.'
login_manager.login_message_category = 'warning'

ALLOWED_EXT = Config.ALLOWED_EXTENSIONS
ALLOWED_DOC_EXT = Config.ALLOWED_DOC_EXTENSIONS


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXT


def allowed_doc(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_DOC_EXT


def _parse_foto_list(lampiran_foto):
    """Parse lampiran_foto field into a list of filenames.

    Handles both legacy single-filename strings and the new JSON-array format.
    Always returns a (possibly empty) list.
    """
    if not lampiran_foto:
        return []
    try:
        parsed = json.loads(lampiran_foto)
        if isinstance(parsed, list):
            return [f for f in parsed if f]
        return [str(parsed)] if parsed else []
    except (json.JSONDecodeError, TypeError):
        return [lampiran_foto] if lampiran_foto else []


def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def wrapper(*a, **kw):
            if current_user.role not in roles:
                abort(403)
            return f(*a, **kw)
        return wrapper
    return decorator


def _q(conn, sql, params=None, one=False):
    """Execute a query and return results as list of dicts (or single dict)."""
    with conn.cursor() as cur:
        cur.execute(sql, params or ())
        if one:
            return cur.fetchone()
        return cur.fetchall()


def _exec(conn, sql, params=None):
    """Execute a write query, commit, and return lastrowid."""
    with conn.cursor() as cur:
        cur.execute(sql, params or ())
    conn.commit()
    return cur.lastrowid


class User(UserMixin):
    def __init__(self, row):
        self.id = row['id']
        self.username = row['username']
        self.nama_lengkap = row['nama_lengkap']
        self.role = row['role']
        self.divisi = row['divisi']

    def get_id(self):
        return str(self.id)


@login_manager.user_loader
def load_user(uid):
    try:
        conn = get_db()
        row = _q(conn, "SELECT * FROM users WHERE id=%s AND is_active=1", (uid,), one=True)
        conn.close()
        return User(row) if row else None
    except Exception:
        return None


@app.teardown_appcontext
def close_db(exc):
    db = g.pop('db', None)
    if db is not None:
        db.close()


@app.context_processor
def inject_globals():
    notif_count = 0
    if current_user.is_authenticated:
        try:
            conn = get_db()
            row = _q(conn, "SELECT COUNT(*) AS c FROM notifications WHERE user_id=%s AND is_read=0",
                     (current_user.id,), one=True)
            notif_count = row['c'] if row else 0
            conn.close()
        except Exception:
            pass
    return dict(
        app_name=Config.APP_NAME,
        company_name=Config.COMPANY_NAME,
        company_sub=Config.COMPANY_SUB,
        now=datetime.now(),
        notif_count=notif_count,
    )


# ---------------------------------------------------------------------------
# Auth routes
# ---------------------------------------------------------------------------
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        conn = get_db()
        row = _q(conn, "SELECT * FROM users WHERE username=%s AND is_active=1",
                 (username,), one=True)
        conn.close()
        if row and bcrypt.checkpw(password.encode(), row['password'].encode()):
            login_user(User(row))
            _log(row['id'], 'LOGIN', f'{username} logged in')
            flash('Login berhasil!', 'success')
            nxt = request.args.get('next')
            if nxt and urlparse(nxt).netloc != '':
                nxt = None
            return redirect(nxt or url_for('dashboard'))
        flash('Username atau password salah.', 'danger')
    return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        nama = request.form.get('nama_lengkap', '').strip()
        divisi = request.form.get('divisi', '').strip()
        if not all([username, password, nama]):
            flash('Semua field wajib diisi.', 'danger')
            return redirect(url_for('register'))
        hashed = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
        conn = get_db()
        try:
            _exec(conn,
                  "INSERT INTO users (username,password,nama_lengkap,role,divisi) VALUES (%s,%s,%s,%s,%s)",
                  (username, hashed, nama, 'staff', divisi))
            flash('Registrasi berhasil! Silakan login.', 'success')
            return redirect(url_for('login'))
        except Exception:
            flash('Username sudah digunakan.', 'danger')
        finally:
            conn.close()
    return render_template('register.html')


@app.route('/logout')
@login_required
def logout():
    _log(current_user.id, 'LOGOUT', f'{current_user.username} logged out')
    logout_user()
    flash('Anda telah logout.', 'info')
    return redirect(url_for('login'))


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------
@app.route('/')
@login_required
def dashboard():
    conn = get_db()
    total_keluar = _q(conn,
        "SELECT COUNT(*) AS c FROM surat_izin WHERE jenis='keluar'", one=True)['c']
    total_masuk = _q(conn,
        "SELECT COUNT(*) AS c FROM surat_izin WHERE jenis='masuk'", one=True)['c']
    total_pending = _q(conn,
        "SELECT COUNT(*) AS c FROM surat_izin WHERE status='pending'", one=True)['c']
    total_review = _q(conn,
        "SELECT COUNT(*) AS c FROM surat_izin WHERE status='review'", one=True)['c']
    total_approved = _q(conn,
        "SELECT COUNT(*) AS c FROM surat_izin WHERE status='approved'", one=True)['c']

    recent = _q(conn,
        "SELECT * FROM surat_izin ORDER BY created_at DESC LIMIT 5")

    monthly = _q(conn, """
        SELECT DATE_FORMAT(tanggal, '%%Y-%%m') AS bulan,
               SUM(CASE WHEN jenis='keluar' THEN 1 ELSE 0 END) AS keluar,
               SUM(CASE WHEN jenis='masuk' THEN 1 ELSE 0 END) AS masuk
        FROM surat_izin
        GROUP BY bulan ORDER BY bulan DESC LIMIT 6
    """)

    divisi_stats = _q(conn, """
        SELECT divisi, COUNT(*) AS total FROM surat_izin GROUP BY divisi ORDER BY total DESC
    """)

    conn.close()
    return render_template('dashboard.html',
                           total_keluar=total_keluar,
                           total_masuk=total_masuk,
                           total_pending=total_pending,
                           total_review=total_review,
                           total_approved=total_approved,
                           recent=recent,
                           monthly=list(reversed(monthly)),
                           divisi_stats=divisi_stats)


# ---------------------------------------------------------------------------
# Surat CRUD
# ---------------------------------------------------------------------------
@app.route('/surat')
@login_required
def surat_list():
    jenis = request.args.get('jenis', '')
    status = request.args.get('status', '')
    search = request.args.get('q', '')

    q = "SELECT * FROM surat_izin WHERE 1=1"
    params = []
    if jenis:
        q += " AND jenis=%s"
        params.append(jenis)
    if status:
        q += " AND status=%s"
        params.append(status)
    if search:
        q += " AND (no_surat LIKE %s OR nama LIKE %s OR perusahaan LIKE %s)"
        params.extend([f'%{search}%'] * 3)
    q += " ORDER BY created_at DESC"

    conn = get_db()
    rows = _q(conn, q, params)
    conn.close()
    return render_template('surat_list.html', surat_list=rows,
                           jenis=jenis, status=status, search=search)


@app.route('/surat/add', methods=['GET', 'POST'])
@login_required
def add_surat():
    if request.method == 'POST':
        jenis = request.form.get('jenis', 'keluar')
        required = ['no_surat', 'tanggal', 'tgl_terbit', 'divisi', 'nama',
                     'badge', 'no_kendaraan', 'perusahaan', 'no_spk',
                     'pemohon', 'diperiksa_oleh', 'disetujui_oleh']
        for f in required:
            if not request.form.get(f):
                flash(f'Field {f} harus diisi!', 'danger')
                return redirect(url_for('add_surat'))

        # Handle KTP photo upload
        foto_ktp_name = None
        ktp_file = request.files.get('foto_ktp')
        if ktp_file and ktp_file.filename and allowed_file(ktp_file.filename):
            ext = ktp_file.filename.rsplit('.', 1)[1].lower()
            foto_ktp_name = f"{uuid.uuid4().hex}.{ext}"
            ktp_file.save(os.path.join(Config.UPLOAD_DIR, foto_ktp_name))

        # Handle SPK file upload
        file_spk_name = None
        spk_file = request.files.get('file_spk')
        if spk_file and spk_file.filename and allowed_doc(spk_file.filename):
            ext = spk_file.filename.rsplit('.', 1)[1].lower()
            file_spk_name = f"{uuid.uuid4().hex}.{ext}"
            spk_file.save(os.path.join(Config.UPLOAD_DIR, file_spk_name))

        try:
            barang = json.loads(request.form.get('barang_items', '[]'))
        except json.JSONDecodeError:
            flash('Format data barang tidak valid!', 'danger')
            return redirect(url_for('add_surat'))

        # Handle per-item photo uploads
        for i, item in enumerate(barang):
            item_fotos = request.files.getlist(f'foto_barang_{i}')
            item_foto_names = []
            for foto in item_fotos:
                if foto and foto.filename and allowed_file(foto.filename):
                    ext = foto.filename.rsplit('.', 1)[1].lower()
                    fname = f"{uuid.uuid4().hex}.{ext}"
                    foto.save(os.path.join(Config.UPLOAD_DIR, fname))
                    item_foto_names.append(fname)
            if item_foto_names:
                item['foto'] = item_foto_names

        conn = get_db()
        sid = _exec(conn, """
            INSERT INTO surat_izin
            (jenis,no_surat,tanggal,tgl_terbit,divisi,nama,badge,no_kendaraan,
             perusahaan,no_spk,foto_ktp,file_spk,pemohon,diperiksa_oleh,disetujui_oleh,
             barang_items,lampiran_foto,status,created_by)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            jenis,
            request.form['no_surat'].strip(),
            request.form['tanggal'],
            request.form['tgl_terbit'],
            request.form['divisi'],
            request.form['nama'].strip(),
            request.form['badge'].strip(),
            request.form['no_kendaraan'].strip(),
            request.form['perusahaan'].strip(),
            request.form['no_spk'].strip(),
            foto_ktp_name,
            file_spk_name,
            request.form['pemohon'].strip(),
            request.form['diperiksa_oleh'].strip(),
            request.form['disetujui_oleh'].strip(),
            json.dumps(barang, ensure_ascii=False),
            None,  # lampiran_foto — photos are now managed per-item in barang_items
            'pending',
            current_user.id,
        ))
        conn.close()
        _log(current_user.id, 'CREATE',
             f'Surat {jenis} {request.form["no_surat"]} dibuat')
        _notify_all_admins(
            'Surat Baru Dibuat',
            f'{current_user.nama_lengkap} membuat surat {jenis} {request.form["no_surat"]}',
            f'/surat/{sid}',
            exclude_user=current_user.id,
        )
        flash('Surat berhasil dibuat!', 'success')
        return redirect(url_for('view_surat', id=sid))

    today = date.today().isoformat()
    jenis = request.args.get('jenis', 'keluar')
    return render_template('add_surat.html', date_now=today, jenis=jenis)


@app.route('/surat/<int:id>')
@login_required
def view_surat(id):
    conn = get_db()
    surat = _q(conn, "SELECT * FROM surat_izin WHERE id=%s", (id,), one=True)
    if not surat:
        conn.close()
        flash('Surat tidak ditemukan.', 'danger')
        return redirect(url_for('surat_list'))
    surat = dict(surat)
    try:
        surat['barang_items'] = json.loads(surat['barang_items'])
    except Exception:
        surat['barang_items'] = []
    surat['foto_list'] = _parse_foto_list(surat.get('lampiran_foto'))
    # Look up approver names for signature section
    for key in ('approval_user_by', 'approval_satpam_by', 'approval_asman_by', 'approval_manager_by'):
        uid = surat.get(key)
        if uid:
            u = _q(conn, "SELECT nama_lengkap FROM users WHERE id=%s", (uid,), one=True)
            surat[key + '_name'] = u['nama_lengkap'] if u else '-'
        else:
            surat[key + '_name'] = ''
    # Also look up creator name
    creator_id = surat.get('created_by')
    if creator_id:
        creator = _q(conn, "SELECT nama_lengkap FROM users WHERE id=%s", (creator_id,), one=True)
        surat['creator_name'] = creator['nama_lengkap'] if creator else '-'
    else:
        surat['creator_name'] = ''
    conn.close()
    return render_template('view_surat.html', surat=surat)


@app.route('/surat/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_surat(id):
    conn = get_db()
    surat = _q(conn, "SELECT * FROM surat_izin WHERE id=%s", (id,), one=True)
    if not surat:
        conn.close()
        flash('Surat tidak ditemukan.', 'danger')
        return redirect(url_for('surat_list'))
    if request.method == 'POST':
        # Handle KTP photo upload
        foto_ktp_name = surat.get('foto_ktp')
        ktp_file = request.files.get('foto_ktp')
        if ktp_file and ktp_file.filename and allowed_file(ktp_file.filename):
            ext = ktp_file.filename.rsplit('.', 1)[1].lower()
            foto_ktp_name = f"{uuid.uuid4().hex}.{ext}"
            ktp_file.save(os.path.join(Config.UPLOAD_DIR, foto_ktp_name))

        # Handle SPK file upload
        file_spk_name = surat.get('file_spk')
        spk_file = request.files.get('file_spk')
        if spk_file and spk_file.filename and allowed_doc(spk_file.filename):
            ext = spk_file.filename.rsplit('.', 1)[1].lower()
            file_spk_name = f"{uuid.uuid4().hex}.{ext}"
            spk_file.save(os.path.join(Config.UPLOAD_DIR, file_spk_name))

        try:
            barang = json.loads(request.form.get('barang_items', '[]'))
        except json.JSONDecodeError:
            flash('Format data barang tidak valid!', 'danger')
            return redirect(url_for('edit_surat', id=id))

        # Handle per-item photo uploads
        for i, item in enumerate(barang):
            item_fotos = request.files.getlist(f'foto_barang_{i}')
            item_foto_names = item.get('foto', [])
            if not isinstance(item_foto_names, list):
                item_foto_names = []
            for foto in item_fotos:
                if foto and foto.filename and allowed_file(foto.filename):
                    ext = foto.filename.rsplit('.', 1)[1].lower()
                    fname = f"{uuid.uuid4().hex}.{ext}"
                    foto.save(os.path.join(Config.UPLOAD_DIR, fname))
                    item_foto_names.append(fname)
            if item_foto_names:
                item['foto'] = item_foto_names

        _exec(conn, """
            UPDATE surat_izin SET
              jenis=%s,no_surat=%s,tanggal=%s,tgl_terbit=%s,divisi=%s,nama=%s,badge=%s,
              no_kendaraan=%s,perusahaan=%s,no_spk=%s,foto_ktp=%s,file_spk=%s,pemohon=%s,
              diperiksa_oleh=%s,disetujui_oleh=%s,barang_items=%s
            WHERE id=%s
        """, (
            request.form.get('jenis', surat['jenis']),
            request.form['no_surat'].strip(),
            request.form['tanggal'],
            request.form['tgl_terbit'],
            request.form['divisi'],
            request.form['nama'].strip(),
            request.form['badge'].strip(),
            request.form['no_kendaraan'].strip(),
            request.form['perusahaan'].strip(),
            request.form['no_spk'].strip(),
            foto_ktp_name,
            file_spk_name,
            request.form['pemohon'].strip(),
            request.form['diperiksa_oleh'].strip(),
            request.form['disetujui_oleh'].strip(),
            json.dumps(barang, ensure_ascii=False),
            id,
        ))
        conn.close()
        _log(current_user.id, 'UPDATE', f'Surat #{id} diperbarui')
        flash('Surat berhasil diperbarui!', 'success')
        return redirect(url_for('view_surat', id=id))

    conn.close()
    surat = dict(surat)
    try:
        surat['barang_items'] = json.loads(surat['barang_items'])
    except Exception:
        surat['barang_items'] = []
    surat['foto_list'] = _parse_foto_list(surat.get('lampiran_foto'))
    return render_template('edit_surat.html', surat=surat)


@app.route('/surat/<int:id>/status', methods=['POST'])
@login_required
@role_required('admin')
def update_status(id):
    new_status = request.form.get('status')
    catatan = request.form.get('catatan', '')
    if new_status not in ('approved', 'rejected', 'pending', 'review'):
        flash('Status tidak valid.', 'danger')
        return redirect(url_for('view_surat', id=id))
    conn = get_db()
    surat = _q(conn, "SELECT no_surat, created_by FROM surat_izin WHERE id=%s", (id,), one=True)
    _exec(conn, "UPDATE surat_izin SET status=%s, catatan=%s WHERE id=%s",
          (new_status, catatan, id))
    conn.close()
    _log(current_user.id, 'STATUS', f'Surat #{id} status → {new_status}')

    # Notify the surat creator about status change
    if surat and surat.get('created_by'):
        status_labels = {'approved': 'Disetujui', 'rejected': 'Ditolak',
                         'review': 'Sedang Direview', 'pending': 'Pending'}
        label = status_labels.get(new_status, new_status)
        _notify(surat['created_by'],
                f'Status Surat Diperbarui',
                f'Surat {surat["no_surat"]} status diubah menjadi {label}',
                f'/surat/{id}')

    flash(f'Status surat diubah menjadi {new_status}.', 'success')
    return redirect(url_for('view_surat', id=id))


@app.route('/surat/<int:id>/approve', methods=['POST'])
@login_required
def approve_surat(id):
    """Multi-stage approval: User -> Satpam -> Asman Umum -> Manager Administrasi"""
    stage = request.form.get('stage')
    decision = request.form.get('decision')
    note = request.form.get('note', '')

    conn = get_db()
    surat = _q(conn, "SELECT * FROM surat_izin WHERE id=%s", (id,), one=True)
    if not surat:
        conn.close()
        flash('Surat tidak ditemukan.', 'danger')
        return redirect(url_for('surat_list'))

    if stage == 'user':
        # User / Pemberi Kerja (user, staff, or admin) — NOT the pemohon/creator
        if current_user.role not in ('user', 'admin'):
            abort(403)
        if decision not in ('sesuai', 'tidak_sesuai'):
            flash('Keputusan tidak valid.', 'danger')
            return redirect(url_for('view_surat', id=id))
        # Save per-item user approvals from form
        try:
            barang_items = json.loads(surat['barang_items']) if surat['barang_items'] else []
        except (json.JSONDecodeError, TypeError):
            barang_items = []
        for i, item in enumerate(barang_items):
            item_approval = request.form.get(f'item_user_approval_{i}', 'pending')
            item['approval_user'] = item_approval
        _exec(conn, """UPDATE surat_izin SET
            approval_user=%s, approval_user_by=%s, approval_user_at=NOW(),
            approval_user_note=%s, barang_items=%s
            WHERE id=%s""",
            (decision, current_user.id, note, json.dumps(barang_items, ensure_ascii=False), id))
        label = 'Sesuai' if decision == 'sesuai' else 'Tidak Sesuai'
        _log(current_user.id, 'APPROVE', f'User/Pemberi Kerja: Surat #{id} → {label}')
        # Notify satpam users
        satpams = _q(conn, "SELECT id FROM users WHERE role IN ('satpam','admin') AND is_active=1")
        for s in satpams:
            if s['id'] != current_user.id:
                _notify(s['id'], 'Perlu Pemeriksaan Satpam', f'Surat {surat["no_surat"]} sudah dicek User/Pemberi Kerja ({label}), perlu pemeriksaan', f'/surat/{id}')
    elif stage == 'satpam':
        if current_user.role not in ('satpam', 'admin'):
            abort(403)
        if surat.get('approval_user') not in ('sesuai', 'tidak_sesuai'):
            flash('User/Pemberi Kerja belum memeriksa surat ini.', 'warning')
            return redirect(url_for('view_surat', id=id))
        if decision not in ('sesuai', 'tidak_sesuai'):
            flash('Keputusan tidak valid.', 'danger')
            return redirect(url_for('view_surat', id=id))
        # Save per-item approvals from form
        try:
            barang_items = json.loads(surat['barang_items']) if surat['barang_items'] else []
        except (json.JSONDecodeError, TypeError):
            barang_items = []
        for i, item in enumerate(barang_items):
            item_approval = request.form.get(f'item_approval_{i}', 'pending')
            item['approval_satpam'] = item_approval
        _exec(conn, """UPDATE surat_izin SET
            approval_satpam=%s, approval_satpam_by=%s, approval_satpam_at=NOW(),
            approval_satpam_note=%s, barang_items=%s, status='review'
            WHERE id=%s""",
            (decision, current_user.id, note, json.dumps(barang_items, ensure_ascii=False), id))
        label = 'Sesuai' if decision == 'sesuai' else 'Tidak Sesuai'
        _log(current_user.id, 'APPROVE', f'Satpam: Surat #{id} \u2192 {label}')
        # Notify asman users
        asmans = _q(conn, "SELECT id FROM users WHERE role IN ('asman','admin') AND is_active=1")
        for a in asmans:
            if a['id'] != current_user.id:
                _notify(a['id'], 'Perlu Review Asman', f'Surat {surat["no_surat"]} sudah diperiksa Satpam ({label})', f'/surat/{id}')
    elif stage == 'asman':
        if current_user.role not in ('asman', 'admin'):
            abort(403)
        if surat.get('approval_satpam') == 'pending':
            flash('Satpam belum memeriksa surat ini.', 'warning')
            return redirect(url_for('view_surat', id=id))
        if decision not in ('approved', 'rejected'):
            flash('Keputusan tidak valid.', 'danger')
            return redirect(url_for('view_surat', id=id))
        _exec(conn, """UPDATE surat_izin SET
            approval_asman=%s, approval_asman_by=%s, approval_asman_at=NOW(),
            approval_asman_note=%s
            WHERE id=%s""",
            (decision, current_user.id, note, id))
        label = 'Disetujui' if decision == 'approved' else 'Ditolak'
        _log(current_user.id, 'APPROVE', f'Asman: Surat #{id} \u2192 {label}')
        # Notify managers
        managers = _q(conn, "SELECT id FROM users WHERE role IN ('manager','admin') AND is_active=1")
        for m in managers:
            if m['id'] != current_user.id:
                _notify(m['id'], 'Perlu Approval Manager', f'Surat {surat["no_surat"]} sudah direview Asman ({label})', f'/surat/{id}')
    elif stage == 'manager':
        if current_user.role not in ('manager', 'admin'):
            abort(403)
        if surat.get('approval_asman') == 'pending':
            flash('Asman belum mereview surat ini.', 'warning')
            return redirect(url_for('view_surat', id=id))
        if decision not in ('approved', 'rejected'):
            flash('Keputusan tidak valid.', 'danger')
            return redirect(url_for('view_surat', id=id))
        final_status = decision
        _exec(conn, """UPDATE surat_izin SET
            approval_manager=%s, approval_manager_by=%s, approval_manager_at=NOW(),
            approval_manager_note=%s, status=%s
            WHERE id=%s""",
            (decision, current_user.id, note, final_status, id))
        label = 'Disetujui' if decision == 'approved' else 'Ditolak'
        _log(current_user.id, 'APPROVE', f'Manager: Surat #{id} \u2192 {label}')
        # Notify creator
        if surat.get('created_by'):
            _notify(surat['created_by'], 'Surat Final', f'Surat {surat["no_surat"]} telah {label} oleh Manager', f'/surat/{id}')
    else:
        flash('Stage tidak valid.', 'danger')
        conn.close()
        return redirect(url_for('view_surat', id=id))

    conn.close()
    flash('Approval berhasil disimpan.', 'success')
    return redirect(url_for('view_surat', id=id))


@app.route('/surat/<int:id>/delete')
@login_required
@role_required('admin', 'manager')
def delete_surat(id):
    conn = get_db()
    s = _q(conn, "SELECT no_surat FROM surat_izin WHERE id=%s", (id,), one=True)
    if s:
        _exec(conn, "DELETE FROM surat_izin WHERE id=%s", (id,))
        _log(current_user.id, 'DELETE', f'Surat {s["no_surat"]} dihapus')
        flash('Surat berhasil dihapus.', 'success')
    else:
        flash('Surat tidak ditemukan.', 'danger')
    conn.close()
    return redirect(url_for('surat_list'))


# ---------------------------------------------------------------------------
# Export PDF (single surat)
# ---------------------------------------------------------------------------
@app.route('/surat/<int:id>/pdf')
@login_required
def export_pdf(id):
    conn = get_db()
    surat = _q(conn, "SELECT * FROM surat_izin WHERE id=%s", (id,), one=True)
    if not surat:
        conn.close()
        flash('Surat tidak ditemukan.', 'danger')
        return redirect(url_for('surat_list'))
    surat = dict(surat)
    try:
        surat['barang_items'] = json.loads(surat['barang_items'])
    except Exception:
        surat['barang_items'] = []
    surat['foto_list'] = _parse_foto_list(surat.get('lampiran_foto'))

    # Look up approver names for the signature section
    for key in ('approval_user_by', 'approval_satpam_by', 'approval_asman_by', 'approval_manager_by'):
        uid = surat.get(key)
        if uid:
            u = _q(conn, "SELECT nama_lengkap FROM users WHERE id=%s", (uid,), one=True)
            surat[key + '_name'] = u['nama_lengkap'] if u else '-'
        else:
            surat[key + '_name'] = ''
    # Look up creator name for Pemohon
    creator_id = surat.get('created_by')
    if creator_id:
        creator = _q(conn, "SELECT nama_lengkap FROM users WHERE id=%s", (creator_id,), one=True)
        surat['creator_name'] = creator['nama_lengkap'] if creator else '-'
    else:
        surat['creator_name'] = ''
    conn.close()

    # Build SPK download URL for the PDF link
    spk_url = None
    if surat.get('file_spk'):
        spk_url = request.host_url.rstrip('/') + url_for('static', filename='uploads/' + surat['file_spk'])

    html = render_template('pdf_template.html', surat=surat,
                           upload_dir=os.path.abspath(Config.UPLOAD_DIR),
                           spk_url=spk_url)
    try:
        from weasyprint import HTML
        from pathlib import Path
        base = Path(app.root_path, 'static').as_uri() + '/'
        pdf_bytes = HTML(string=html, base_url=base).write_pdf()
        clean = ''.join(c for c in surat['no_surat'] if c.isalnum() or c in '-_.')
        filename = f"surat_izin_{clean}.pdf"
        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        flash(f'Gagal membuat PDF: {e}', 'danger')
        return redirect(url_for('view_surat', id=id))


# ---------------------------------------------------------------------------
# Export Excel report
# ---------------------------------------------------------------------------
@app.route('/report/excel')
@login_required
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    conn = get_db()
    rows = _q(conn, "SELECT * FROM surat_izin ORDER BY tanggal DESC")
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Surat Izin"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    ws.merge_cells('A1:I1')
    ws['A1'] = f'LAPORAN SURAT IZIN KELUAR MASUK BARANG - {Config.COMPANY_NAME}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:I2')
    ws['A2'] = f'Dicetak: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['No', 'Jenis', 'No. Surat', 'Tanggal', 'Divisi',
               'Nama', 'Perusahaan', 'Status', 'Dibuat']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, r in enumerate(rows, 5):
        vals = [i - 4, r['jenis'].upper(), r['no_surat'], str(r['tanggal']),
                r['divisi'], r['nama'], r['perusahaan'],
                r['status'].upper(), str(r['created_at'])]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center' if col in (1, 2, 8) else 'left')

    from openpyxl.utils import get_column_letter
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 10
        for row_cells in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=4):
            for cell in row_cells:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'laporan_surat_{datetime.now().strftime("%Y%m%d")}.xlsx')


# ---------------------------------------------------------------------------
# Export PDF report
# ---------------------------------------------------------------------------
@app.route('/report/pdf')
@login_required
def export_report_pdf():
    conn = get_db()
    rows = _q(conn, "SELECT * FROM surat_izin ORDER BY tanggal DESC")
    total_keluar = _q(conn, "SELECT COUNT(*) AS c FROM surat_izin WHERE jenis='keluar'", one=True)['c']
    total_masuk = _q(conn, "SELECT COUNT(*) AS c FROM surat_izin WHERE jenis='masuk'", one=True)['c']
    conn.close()

    html = render_template('report_pdf.html', rows=rows,
                           total_keluar=total_keluar, total_masuk=total_masuk)
    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=html).write_pdf()
        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype='application/pdf', as_attachment=True,
            download_name=f'laporan_{datetime.now().strftime("%Y%m%d")}.pdf',
        )
    except Exception as e:
        flash(f'Gagal membuat PDF: {e}', 'danger')
        return redirect(url_for('dashboard'))


# ---------------------------------------------------------------------------
# Barang Report (with filtering, search, PDF & Excel export)
# ---------------------------------------------------------------------------
def _build_barang_query(params):
    """Build SQL query and param list from filter request args."""
    jenis = params.get('jenis', '')
    status = params.get('status', '')
    date_from = params.get('date_from', '')
    date_to = params.get('date_to', '')
    search = params.get('q', '')

    q = "SELECT * FROM surat_izin WHERE 1=1"
    p = []
    if jenis:
        q += " AND jenis=%s"; p.append(jenis)
    if status:
        q += " AND status=%s"; p.append(status)
    if date_from:
        q += " AND tanggal >= %s"; p.append(date_from)
    if date_to:
        q += " AND tanggal <= %s"; p.append(date_to)
    if search:
        q += " AND (no_surat LIKE %s OR nama LIKE %s OR perusahaan LIKE %s)"
        p.extend([f'%{search}%'] * 3)
    q += " ORDER BY tanggal DESC"
    return q, p


def _flatten_barang(rows):
    """Flatten surat rows into per-item barang records for reporting."""
    items = []
    for r in rows:
        try:
            barang = json.loads(r['barang_items']) if r['barang_items'] else []
        except (json.JSONDecodeError, TypeError):
            barang = []
        for b in barang:
            items.append({
                'surat_id': r['id'],
                'jenis': r['jenis'],
                'no_surat': r['no_surat'],
                'tanggal': r['tanggal'],
                'divisi': r['divisi'],
                'nama_pemohon': r['nama'],
                'perusahaan': r['perusahaan'],
                'status': r['status'],
                'nama_barang': b.get('nama_barang', ''),
                'jumlah': b.get('jumlah', ''),
                'satuan': b.get('satuan', ''),
                'keterangan': b.get('keterangan', ''),
                'foto': b.get('foto', []),
                'approval_user': b.get('approval_user', ''),
                'approval_satpam': b.get('approval_satpam', ''),
            })
    return items


@app.route('/report/barang')
@login_required
def report_barang():
    conn = get_db()
    q, p = _build_barang_query(request.args)
    rows = _q(conn, q, p)
    conn.close()
    items = _flatten_barang(rows)

    total_keluar = sum(1 for r in rows if r['jenis'] == 'keluar')
    total_masuk = sum(1 for r in rows if r['jenis'] == 'masuk')

    return render_template('report_barang.html', items=items,
                           total_keluar=total_keluar, total_masuk=total_masuk,
                           total_barang=len(items),
                           jenis=request.args.get('jenis', ''),
                           status=request.args.get('status', ''),
                           date_from=request.args.get('date_from', ''),
                           date_to=request.args.get('date_to', ''),
                           search=request.args.get('q', ''))


@app.route('/report/barang/excel')
@login_required
def report_barang_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.utils import get_column_letter

    conn = get_db()
    q, p = _build_barang_query(request.args)
    rows = _q(conn, q, p)
    conn.close()
    items = _flatten_barang(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Barang"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    ws.merge_cells('A1:M1')
    ws['A1'] = f'LAPORAN BARANG MASUK & KELUAR - {Config.COMPANY_NAME}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:M2')
    filters_text = f'Dicetak: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    jenis_f = request.args.get('jenis', '')
    status_f = request.args.get('status', '')
    if jenis_f:
        filters_text += f' | Jenis: {jenis_f.upper()}'
    if status_f:
        filters_text += f' | Status: {status_f.upper()}'
    ws['A2'] = filters_text
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['No', 'Jenis', 'No. Surat', 'Tanggal', 'Divisi', 'Pemohon',
               'Perusahaan', 'Nama Barang', 'Jumlah', 'Satuan', 'Keterangan',
               'Cek User', 'Cek Satpam']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, item in enumerate(items, 5):
        user_approval_text = ''
        if item['approval_user'] == 'sesuai':
            user_approval_text = '✓ Sesuai'
        elif item['approval_user'] == 'tidak_sesuai':
            user_approval_text = '✗ Tidak Sesuai'

        approval_text = ''
        if item['approval_satpam'] == 'sesuai':
            approval_text = '✓ Sesuai'
        elif item['approval_satpam'] == 'tidak_sesuai':
            approval_text = '✗ Tidak Sesuai'

        vals = [i - 4, item['jenis'].upper(), item['no_surat'],
                str(item['tanggal']), item['divisi'], item['nama_pemohon'],
                item['perusahaan'], item['nama_barang'],
                item['jumlah'], item['satuan'],
                item['keterangan'] or '-', user_approval_text, approval_text]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center' if col in (1, 2, 9, 10, 12, 13) else 'left')
            if col == 12 and item['approval_user'] == 'sesuai':
                cell.fill = green_fill
            elif col == 12 and item['approval_user'] == 'tidak_sesuai':
                cell.fill = red_fill
            if col == 13 and item['approval_satpam'] == 'sesuai':
                cell.fill = green_fill
            elif col == 13 and item['approval_satpam'] == 'tidak_sesuai':
                cell.fill = red_fill

        # Add photos in column N onward
        if item['foto']:
            foto_col = len(headers) + 1
            for foto_name in item['foto']:
                foto_path = os.path.join(Config.UPLOAD_DIR, foto_name)
                if os.path.exists(foto_path):
                    try:
                        img = XlImage(foto_path)
                        img.width = 60
                        img.height = 60
                        cell_ref = f'{get_column_letter(foto_col)}{i}'
                        ws.add_image(img, cell_ref)
                        ws.row_dimensions[i].height = 50
                        foto_col += 1
                    except Exception:
                        pass

    # Auto-size columns
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 10
        for row_cells in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=4):
            for cell in row_cells:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'laporan_barang_{datetime.now().strftime("%Y%m%d")}.xlsx')


@app.route('/report/barang/pdf')
@login_required
def report_barang_pdf():
    conn = get_db()
    q, p = _build_barang_query(request.args)
    rows = _q(conn, q, p)
    conn.close()
    items = _flatten_barang(rows)

    total_keluar = sum(1 for r in rows if r['jenis'] == 'keluar')
    total_masuk = sum(1 for r in rows if r['jenis'] == 'masuk')

    html = render_template('report_barang_pdf.html', items=items,
                           total_keluar=total_keluar, total_masuk=total_masuk,
                           total_barang=len(items),
                           upload_dir=os.path.abspath(Config.UPLOAD_DIR))
    try:
        from weasyprint import HTML
        from pathlib import Path
        base = Path(app.root_path, 'static').as_uri() + '/'
        pdf_bytes = HTML(string=html, base_url=base).write_pdf()
        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype='application/pdf', as_attachment=True,
            download_name=f'laporan_barang_{datetime.now().strftime("%Y%m%d")}.pdf',
        )
    except Exception as e:
        flash(f'Gagal membuat PDF: {e}', 'danger')
        return redirect(url_for('report_barang'))


# ---------------------------------------------------------------------------
# User management (admin only)
# ---------------------------------------------------------------------------
@app.route('/users')
@login_required
@role_required('admin')
def user_list():
    conn = get_db()
    users = _q(conn, "SELECT * FROM users ORDER BY created_at DESC")
    conn.close()
    return render_template('users.html', users=users)


@app.route('/users/<int:uid>/toggle')
@login_required
@role_required('admin')
def toggle_user(uid):
    if uid == current_user.id:
        flash('Tidak bisa menonaktifkan diri sendiri.', 'danger')
        return redirect(url_for('user_list'))
    conn = get_db()
    u = _q(conn, "SELECT is_active FROM users WHERE id=%s", (uid,), one=True)
    if u:
        _exec(conn, "UPDATE users SET is_active=%s WHERE id=%s",
              (0 if u['is_active'] else 1, uid))
        flash('Status user diperbarui.', 'success')
    conn.close()
    return redirect(url_for('user_list'))


@app.route('/users/<int:uid>/role', methods=['POST'])
@login_required
@role_required('admin')
def change_role(uid):
    new_role = request.form.get('role')
    if new_role not in ('admin', 'user', 'staff', 'manager', 'satpam', 'asman'):
        flash('Role tidak valid.', 'danger')
        return redirect(url_for('user_list'))
    conn = get_db()
    _exec(conn, "UPDATE users SET role=%s WHERE id=%s", (new_role, uid))
    conn.close()
    flash('Role user diperbarui.', 'success')
    return redirect(url_for('user_list'))


# ---------------------------------------------------------------------------
# API
# ---------------------------------------------------------------------------
@app.route('/api/surat/<int:id>')
@login_required
def api_surat(id):
    conn = get_db()
    s = _q(conn, "SELECT * FROM surat_izin WHERE id=%s", (id,), one=True)
    conn.close()
    if not s:
        return jsonify(success=False, error='Tidak ditemukan'), 404
    d = dict(s)
    d['barang_items'] = json.loads(d['barang_items'])
    d['tanggal'] = str(d['tanggal'])
    d['tgl_terbit'] = str(d['tgl_terbit'])
    d['created_at'] = str(d['created_at']) if d['created_at'] else None
    d['updated_at'] = str(d['updated_at']) if d['updated_at'] else None
    return jsonify(success=True, data=d)


@app.route('/api/stats')
@login_required
def api_stats():
    conn = get_db()
    monthly = _q(conn, """
        SELECT DATE_FORMAT(tanggal, '%%Y-%%m') AS bulan,
               SUM(CASE WHEN jenis='keluar' THEN 1 ELSE 0 END) AS keluar,
               SUM(CASE WHEN jenis='masuk' THEN 1 ELSE 0 END) AS masuk
        FROM surat_izin GROUP BY bulan ORDER BY bulan DESC LIMIT 12
    """)
    conn.close()
    return jsonify(monthly)


# ---------------------------------------------------------------------------
# Activity log
# ---------------------------------------------------------------------------
@app.route('/activity')
@login_required
@role_required('admin')
def activity_log():
    conn = get_db()
    logs = _q(conn, """
        SELECT l.*, u.username FROM log_activity l
        LEFT JOIN users u ON l.user_id=u.id
        ORDER BY l.created_at DESC LIMIT 100
    """)
    conn.close()
    return render_template('activity.html', logs=logs)


def _log(user_id, action, desc):
    try:
        conn = get_db()
        _exec(conn,
              "INSERT INTO log_activity (user_id,action,description,ip_address) VALUES (%s,%s,%s,%s)",
              (user_id, action, desc, request.remote_addr))
        conn.close()
    except Exception:
        pass


def _notify(user_id, title, message, link=None):
    """Create a notification for a user."""
    try:
        conn = get_db()
        _exec(conn,
              "INSERT INTO notifications (user_id,title,message,link) VALUES (%s,%s,%s,%s)",
              (user_id, title, message, link))
        conn.close()
    except Exception:
        pass


def _notify_all_admins(title, message, link=None, exclude_user=None):
    """Send notification to all admin and manager users."""
    try:
        conn = get_db()
        admins = _q(conn, "SELECT id FROM users WHERE role IN ('admin','manager','satpam','asman') AND is_active=1")
        for a in admins:
            if exclude_user and a['id'] == exclude_user:
                continue
            _exec(conn,
                  "INSERT INTO notifications (user_id,title,message,link) VALUES (%s,%s,%s,%s)",
                  (a['id'], title, message, link))
        conn.close()
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Notifications
# ---------------------------------------------------------------------------
@app.route('/notifications')
@login_required
def notifications_page():
    conn = get_db()
    notifs = _q(conn,
        "SELECT * FROM notifications WHERE user_id=%s ORDER BY created_at DESC LIMIT 50",
        (current_user.id,))
    conn.close()
    return render_template('notifications.html', notifications=notifs)


@app.route('/api/notifications')
@login_required
def api_notifications():
    conn = get_db()
    notifs = _q(conn,
        "SELECT * FROM notifications WHERE user_id=%s ORDER BY created_at DESC LIMIT 20",
        (current_user.id,))
    conn.close()
    result = []
    for n in notifs:
        result.append({
            'id': n['id'],
            'title': n['title'],
            'message': n['message'],
            'link': n['link'],
            'is_read': n['is_read'],
            'created_at': str(n['created_at']) if n['created_at'] else None,
        })
    return jsonify(result)


@app.route('/api/notifications/read', methods=['POST'])
@login_required
def mark_notifications_read():
    conn = get_db()
    _exec(conn, "UPDATE notifications SET is_read=1 WHERE user_id=%s AND is_read=0",
          (current_user.id,))
    conn.close()
    return jsonify(success=True)


@app.route('/api/notifications/<int:nid>/read', methods=['POST'])
@login_required
def mark_notification_read(nid):
    conn = get_db()
    _exec(conn, "UPDATE notifications SET is_read=1 WHERE id=%s AND user_id=%s",
          (nid, current_user.id))
    conn.close()
    return jsonify(success=True)


# ---------------------------------------------------------------------------
# Error handlers
# ---------------------------------------------------------------------------
@app.errorhandler(403)
def forbidden(e):
    return render_template('error.html', code=403,
                           message='Anda tidak memiliki akses ke halaman ini.'), 403


@app.errorhandler(404)
def not_found(e):
    return render_template('error.html', code=404,
                           message='Halaman tidak ditemukan.'), 404


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == '__main__':
    Config.init_app(app)
    init_db()
    print(f"🚀 {Config.APP_NAME}")
    print("📄 Open http://localhost:8000")
    app.run(debug=True, host='0.0.0.0', port=8000)